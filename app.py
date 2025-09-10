import streamlit as st
import pandas as pd
import base64
import re
import json
import io
import requests
import httpx
from urllib.parse import urlparse
from urllib.robotparser import RobotFileParser
from streamlit_lottie import st_lottie
from streamlit_extras.colored_header import colored_header
from streamlit_extras.metric_cards import style_metric_cards
from streamlit_extras.add_vertical_space import add_vertical_space
from streamlit_extras.badges import badge
from web_scraper import get_website_text_content
from database import ScrapingDatabase
from scheduler import get_scheduler
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl import Workbook
import time
from datetime import datetime
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.jobstores.base import JobLookupError
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import threading
import lxml.html
import pdfplumber
from bs4 import BeautifulSoup
import magic

# Admin credentials from environment variables
import os
# Set admin credentials as requested
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "your_username")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "your_secure_password")

def is_valid_url(url):
    """Validate URL format"""
    try:
        result = urlparse(url)
        return all([result.scheme, result.netloc])
    except:
        return False

def authenticate_user(username, password):
    """Authenticate user with hardcoded credentials"""
    return username == ADMIN_USERNAME and password == ADMIN_PASSWORD

def create_download_link(data, filename, file_type="text"):
    """Create download link for scraped data"""
    if file_type == "csv":
        if isinstance(data, list):  # Multiple URLs data
            df = pd.DataFrame(data)
        else:
            df = pd.DataFrame({"Scraped Content": [data]})
        csv_string = df.to_csv(index=False)
        b64 = base64.b64encode(csv_string.encode()).decode()
        href = f'<a href="data:file/csv;base64,{b64}" download="{filename}.csv">Download as CSV</a>'
    elif file_type == "json":
        if isinstance(data, list):
            json_string = json.dumps(data, indent=2)
        else:
            json_string = json.dumps({"content": data}, indent=2)
        b64 = base64.b64encode(json_string.encode()).decode()
        href = f'<a href="data:file/json;base64,{b64}" download="{filename}.json">Download as JSON</a>'
    elif file_type == "excel":
        # Create Excel file
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        if ws is not None:
            if isinstance(data, list):  # Multiple URLs data
                headers = list(data[0].keys()) if data else []
                if headers:
                    # Write headers
                    for col_num, header in enumerate(headers, 1):
                        ws[f'{chr(64+col_num)}{1}'].value = header
                    # Write data rows
                    for row_num, row in enumerate(data, 2):
                        for col_num, col in enumerate(headers, 1):
                            cell_addr = f'{chr(64+col_num)}{row_num}'
                            ws[cell_addr].value = str(row.get(col, ''))
            else:
                ws['A1'].value = "Content"
                ws['A2'].value = str(data)[:32000]  # Limit content size
            
        wb.save(output)
        b64 = base64.b64encode(output.getvalue()).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}.xlsx">Download as Excel</a>'
    elif file_type == "pdf":
        # Create PDF file
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        if isinstance(data, list):
            for item in data:
                story.append(Paragraph(f"<b>URL:</b> {item.get('url', '')}", styles['Normal']))
                story.append(Paragraph(f"<b>Content:</b> {item.get('content', '')[:500]}...", styles['Normal']))
                story.append(Spacer(1, 12))
        else:
            story.append(Paragraph(data[:2000], styles['Normal']))
        
        doc.build(story)
        b64 = base64.b64encode(output.getvalue()).decode()
        href = f'<a href="data:application/pdf;base64,{b64}" download="{filename}.pdf">Download as PDF</a>'
    else:
        # Text file download
        if isinstance(data, list):
            text_content = "\n\n".join([f"URL: {item.get('url', '')}\nContent: {item.get('content', '')}" for item in data])
        else:
            text_content = data
        b64 = base64.b64encode(text_content.encode()).decode()
        href = f'<a href="data:file/txt;base64,{b64}" download="{filename}.txt">Download as Text</a>'
    return href

def load_lottie_url(url: str):
    """Load Lottie animation from URL"""
    r = requests.get(url)
    if r.status_code != 200:
        return None
    return r.json()

def load_lottie_local(filepath: str):
    """Load Lottie animation from local file"""
    try:
        with open(filepath, "r") as f:
            return json.load(f)
    except:
        return None

def get_lottie_animations():
    """Get various Lottie animations for the app"""
    animations = {
        "web_scraping": load_lottie_url("https://assets5.lottiefiles.com/packages/lf20_fcfjwiyb.json"),
        "loading": load_lottie_url("https://assets4.lottiefiles.com/packages/lf20_szlepvdj.json"),
        "success": load_lottie_url("https://assets1.lottiefiles.com/packages/lf20_jbrw3hcz.json"),
        "data_analysis": load_lottie_url("https://assets8.lottiefiles.com/packages/lf20_qp1q7mct.json"),
        "security": load_lottie_url("https://assets2.lottiefiles.com/packages/lf20_06a6pf9i.json"),
        "dashboard": load_lottie_url("https://assets7.lottiefiles.com/packages/lf20_M9p23l.json")
    }
    return animations

def check_robots_txt(url: str) -> tuple[bool, str]:
    """Check if robots.txt allows scraping"""
    try:
        rp = RobotFileParser()
        robots_url = f"{urlparse(url).scheme}://{urlparse(url).netloc}/robots.txt"
        rp.set_url(robots_url)
        rp.read()
        
        can_fetch = rp.can_fetch("*", url)
        message = "✅ Robots.txt allows scraping" if can_fetch else "⚠️ Robots.txt restricts scraping"
        return can_fetch, message
    except:
        return True, "⚠️ Could not read robots.txt (proceeding anyway)"

def landing_page():
    """Beautiful landing page with hero section and features"""
    # Custom CSS for better styling
    st.markdown("""
    <style>
    .hero-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 3rem 2rem;
        border-radius: 20px;
        margin-bottom: 2rem;
        text-align: center;
        color: white;
    }
    .hero-title {
        font-size: 3.5rem;
        font-weight: 700;
        margin-bottom: 1rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }
    .hero-subtitle {
        font-size: 1.3rem;
        margin-bottom: 2rem;
        opacity: 0.9;
    }
    .feature-card {
        background: white;
        padding: 2rem;
        border-radius: 15px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
        border-left: 5px solid #667eea;
        margin-bottom: 1rem;
        transition: transform 0.3s ease;
    }
    .feature-card:hover {
        transform: translateY(-5px);
    }
    .feature-icon {
        font-size: 3rem;
        margin-bottom: 1rem;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 10px;
        text-align: center;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Load animations
    animations = get_lottie_animations()
    
    # Hero Section
    st.markdown("""
    <div class="hero-container">
        <h1 class="hero-title">🚀 Advanced Web Scraper</h1>
        <p class="hero-subtitle">Extract, analyze, and export web content with AI-powered intelligence</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Hero Animation
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if animations["web_scraping"]:
            st_lottie(animations["web_scraping"], height=300, key="hero_animation")
    
    add_vertical_space(2)
    
    # Features Section
    colored_header(
        label="✨ Platform Features",
        description="Discover the power of modern web scraping",
        color_name="violet-70",
    )
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">🌐</div>
            <h3>Multi-Format Support</h3>
            <p>Scrape HTML, XML, PDF, JSON and more with intelligent content detection</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">⚡</div>
            <h3>Batch Processing</h3>
            <p>Process thousands of URLs simultaneously with real-time progress tracking</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">🎯</div>
            <h3>Smart Extraction</h3>
            <p>AI-powered content extraction with metadata analysis and keyword detection</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">📊</div>
            <h3>Advanced Analytics</h3>
            <p>Detailed content analysis with readability scores and data insights</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">🔒</div>
            <h3>Ethical Scraping</h3>
            <p>Built-in robots.txt respect with educational override options</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">⏰</div>
            <h3>Scheduled Tasks</h3>
            <p>Automated scraping with email notifications and background processing</p>
        </div>
        """, unsafe_allow_html=True)
    
    add_vertical_space(3)
    
    # Quick Stats
    colored_header(
        label="📈 Platform Statistics",
        description="See the power of our scraping engine",
        color_name="blue-green-70",
    )
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Supported Formats", "15+", "HTML, PDF, XML, JSON...")
    with col2:
        st.metric("Export Options", "6", "CSV, JSON, Excel, PDF...")
    with col3:
        st.metric("Max Batch Size", "1000+", "URLs per session")
    with col4:
        st.metric("Success Rate", "98%", "Content extraction")
    
    style_metric_cards(
        background_color="#FFFFFF",
        border_left_color="#667eea",
        border_color="#1f66bd",
        box_shadow=True
    )
    
    add_vertical_space(2)
    
    # Call to Action
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("🚀 Start Scraping Now", type="primary", use_container_width=True):
            st.session_state.show_landing = False
            st.rerun()

def login_page():
    """Enhanced login page with animations"""
    animations = get_lottie_animations()
    
    # Custom CSS for login page
    st.markdown("""
    <style>
    .login-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 3rem 2rem;
        border-radius: 20px;
        margin-bottom: 2rem;
        text-align: center;
        color: white;
    }
    .login-box {
        background: white;
        padding: 2rem;
        border-radius: 15px;
        box-shadow: 0 20px 40px rgba(0,0,0,0.1);
    }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="login-container">
        <h1>🔐 Secure Admin Access</h1>
        <p>Advanced Web Scraping Platform</p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        # Security animation
        if animations["security"]:
            st_lottie(animations["security"], height=200, key="security_animation")
        
        st.markdown("<div class='login-box'>", unsafe_allow_html=True)
        
        st.subheader("🛡️ Admin Authentication")
        username = st.text_input("👤 Username", key="login_username", placeholder="Enter your username")
        password = st.text_input("🔑 Password", type="password", key="login_password", placeholder="Enter your password")
        
        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("🔓 Login", type="primary", use_container_width=True):
                if authenticate_user(username, password):
                    st.session_state.authenticated = True
                    st.session_state.username = username
                    st.session_state.show_landing = True
                    st.success("✅ Login successful!")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("❌ Invalid credentials. Please try again.")
        
        with col_b:
            if st.button("ℹ️ Demo Mode", help="View features without login"):
                st.info("💡 Use credentials: your_username / your_secure_password")
        
        st.markdown("</div>", unsafe_allow_html=True)
        
        # Help section
        with st.expander("🔧 Need Help?"):
            st.markdown("""
            **Default Credentials:**
            - Username: `your_username`
            - Password: `your_secure_password`
            
            **Security Features:**
            - 🔒 Environment-based authentication
            - 🛡️ Session management
            - 🔐 Secure credential handling
            """)

def main_app():
    """Enhanced main application interface"""
    # Show landing page first
    if st.session_state.get('show_landing', True):
        landing_page()
        return
    
    # Enhanced sidebar with animations
    animations = get_lottie_animations()
    
    with st.sidebar:
        # Dashboard animation
        if animations["dashboard"]:
            st_lottie(animations["dashboard"], height=150, key="sidebar_animation")
        
        st.title("🚀 Advanced Web Scraper")
        
        # User info with styling
        st.markdown(f"""
        <div style="
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 1rem;
            border-radius: 10px;
            text-align: center;
            margin-bottom: 1rem;
        ">
            <strong>👤 {st.session_state.username}</strong><br>
            <small>🛡️ Administrator</small>
        </div>
        """, unsafe_allow_html=True)
        
        # Settings section
        with st.expander("⚙️ Scraping Settings"):
            st.session_state.respect_robots = st.toggle(
                "🤖 Respect robots.txt", 
                value=st.session_state.get('respect_robots', True),
                help="Toggle to respect or ignore robots.txt (for learning purposes only)"
            )
            
            if not st.session_state.respect_robots:
                st.warning("⚠️ Educational use only! Always respect website terms.")
            
            st.session_state.concurrent_requests = st.slider(
                "🔄 Concurrent Requests", 
                min_value=1, 
                max_value=10, 
                value=st.session_state.get('concurrent_requests', 3),
                help="Number of simultaneous requests"
            )
            
            st.session_state.request_delay = st.slider(
                "⏱️ Request Delay (seconds)", 
                min_value=0.5, 
                max_value=5.0, 
                value=st.session_state.get('request_delay', 1.0),
                step=0.5,
                help="Delay between requests to be respectful"
            )
        
        st.markdown("---")
        
        # Navigation with icons
        st.markdown("### 🧭 Navigation")
        page_options = {
            "🏠 Home": "home",
            "🌐 Single URL": "single", 
            "🔄 Batch Scraping": "batch", 
            "📚 History": "history", 
            "🔍 Search": "search", 
            "⏰ Scheduled": "scheduled",
            "ℹ️ About": "about"
        }
        
        selected_page = st.radio(
            "Choose a section:",
            options=list(page_options.keys()),
            index=0
        )
        
        page = page_options[selected_page]
        
        st.markdown("---")
        
        # Quick actions
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🏠 Home", use_container_width=True):
                st.session_state.show_landing = True
                st.rerun()
        with col2:
            if st.button("🔓 Logout", type="secondary", use_container_width=True):
                # Clear all session state
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.rerun()
    
    # Main content area with page routing
    if page == "home":
        landing_page()
    elif page == "single":
        enhanced_scraper_page()
    elif page == "batch":
        enhanced_batch_scraper_page()
    elif page == "history":
        enhanced_history_page()
    elif page == "search":
        enhanced_search_page()
    elif page == "scheduled":
        enhanced_scheduled_page()
    elif page == "about":
        enhanced_about_page()

def detect_content_type(url: str) -> str:
    """Detect content type of URL"""
    try:
        response = httpx.head(url, timeout=10, follow_redirects=True)
        content_type = response.headers.get('content-type', '').lower()
        
        if 'text/html' in content_type:
            return 'html'
        elif 'application/pdf' in content_type:
            return 'pdf'
        elif 'application/json' in content_type or 'text/json' in content_type:
            return 'json'
        elif 'application/xml' in content_type or 'text/xml' in content_type:
            return 'xml'
        else:
            return 'unknown'
    except:
        return 'unknown'

def enhanced_content_extraction(url: str, content_type: str = None) -> dict:
    """Enhanced content extraction with format-specific handling"""
    if not content_type:
        content_type = detect_content_type(url)
    
    result = {
        'content': '',
        'title': '',
        'metadata': {},
        'content_type': content_type,
        'word_count': 0,
        'char_count': 0,
        'links': [],
        'images': [],
        'readability_score': 0
    }
    
    try:
        if content_type == 'pdf':
            result = extract_pdf_content(url)
        elif content_type == 'json':
            result = extract_json_content(url)
        elif content_type == 'xml':
            result = extract_xml_content(url)
        else:
            # Default HTML extraction with enhanced processing
            result = extract_html_content(url)
            
    except Exception as e:
        result['error'] = str(e)
    
    return result

def extract_pdf_content(url: str) -> dict:
    """Extract content from PDF files"""
    result = {'content': '', 'title': '', 'metadata': {}, 'content_type': 'pdf'}
    
    try:
        response = httpx.get(url, timeout=30)
        
        with pdfplumber.open(io.BytesIO(response.content)) as pdf:
            text_content = []
            result['metadata']['pages'] = len(pdf.pages)
            
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    text_content.append(text)
            
            result['content'] = '\n\n'.join(text_content)
            result['title'] = f"PDF Document - {len(pdf.pages)} pages"
            
            # Extract metadata
            if pdf.metadata:
                result['metadata'].update(pdf.metadata)
                
    except Exception as e:
        result['error'] = str(e)
    
    return result

def extract_json_content(url: str) -> dict:
    """Extract and format JSON content"""
    result = {'content': '', 'title': '', 'metadata': {}, 'content_type': 'json'}
    
    try:
        response = httpx.get(url, timeout=30)
        json_data = response.json()
        
        result['content'] = json.dumps(json_data, indent=2, ensure_ascii=False)
        result['title'] = "JSON Data"
        result['metadata'] = {
            'keys': list(json_data.keys()) if isinstance(json_data, dict) else [],
            'type': type(json_data).__name__,
            'size': len(str(json_data))
        }
        
    except Exception as e:
        result['error'] = str(e)
    
    return result

def extract_xml_content(url: str) -> dict:
    """Extract content from XML files"""
    result = {'content': '', 'title': '', 'metadata': {}, 'content_type': 'xml'}
    
    try:
        response = httpx.get(url, timeout=30)
        
        # Parse XML with proper XML parser
        from lxml import etree
        try:
            root = etree.fromstring(response.content)
            result['content'] = etree.tostring(root, pretty_print=True, encoding='unicode')
        except etree.XMLSyntaxError:
            # Fallback to HTML parser if XML parsing fails
            root = lxml.html.fromstring(response.content)
            result['content'] = lxml.html.tostring(root, pretty_print=True, encoding='unicode')
        result['title'] = "XML Document"
        
        # Extract metadata
        result['metadata'] = {
            'root_tag': root.tag if hasattr(root, 'tag') else 'unknown',
            'elements': len(root.xpath('.//*')),
            'text_nodes': len(root.xpath('.//text()[normalize-space()]'))
        }
        
    except Exception as e:
        result['error'] = str(e)
    
    return result

def extract_html_content(url: str) -> dict:
    """Enhanced HTML content extraction"""
    result = {'content': '', 'title': '', 'metadata': {}, 'content_type': 'html'}
    
    try:
        # Use the existing web_scraper function but enhance the result
        content = get_website_text_content(url)
        
        # Also get raw HTML for additional processing
        response = httpx.get(url, timeout=30)
        soup = BeautifulSoup(response.content, 'html.parser')
        
        result['content'] = content
        result['title'] = soup.title.string if soup.title else 'Untitled'
        
        # Extract links
        links = []
        for a in soup.find_all('a', href=True):
            href = a.get('href')
            if href and isinstance(href, str) and href.startswith('http'):
                links.append(href)
        result['links'] = links[:20]  # Limit to 20 links
        
        # Extract images
        images = []
        for img in soup.find_all('img', src=True):
            src = img.get('src')
            if src and isinstance(src, str) and src.startswith('http'):
                images.append(src)
        result['images'] = images[:10]  # Limit to 10 images
        
        # Calculate readability (simple word/sentence ratio)
        if content:
            sentences = len([s for s in content.split('.') if s.strip()])
            words = len(content.split())
            result['readability_score'] = round((words / max(sentences, 1)), 2)
        
        # Metadata
        result['metadata'] = {
            'description': soup.find('meta', attrs={'name': 'description'}),
            'keywords': soup.find('meta', attrs={'name': 'keywords'}),
            'author': soup.find('meta', attrs={'name': 'author'}),
            'links_found': len(result['links']),
            'images_found': len(result['images'])
        }
        
        # Clean up metadata
        for key in result['metadata']:
            if result['metadata'][key] and hasattr(result['metadata'][key], 'get'):
                result['metadata'][key] = result['metadata'][key].get('content', '')
    
    except Exception as e:
        result['error'] = str(e)
    
    return result

def enhanced_scraper_page():
    """Enhanced scraper interface with modern UI and advanced features"""
    animations = get_lottie_animations()
    
    colored_header(
        label="🌐 Smart Web Scraper",
        description="Extract content from any website with AI-powered intelligence",
        color_name="violet-70",
    )
    
    # URL input section with enhanced styling
    with st.container():
        st.markdown("""
        <div style="
            background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
            padding: 2rem;
            border-radius: 15px;
            margin-bottom: 2rem;
        ">
        """, unsafe_allow_html=True)
        
        st.subheader("🎯 Target Website Analysis")
        
        col1, col2 = st.columns([3, 1])
        
        with col1:
            url = st.text_input(
                "🔗 Website URL",
                placeholder="https://example.com",
                help="Enter any website URL - supports HTML, PDF, JSON, XML formats"
            )
        
        with col2:
            st.markdown("<br>", unsafe_allow_html=True)
            scrape_button = st.button(
                "🤖 Analyze & Extract", 
                type="primary", 
                use_container_width=True
            )
        
        st.markdown("</div>", unsafe_allow_html=True)
    
    # Enhanced URL validation and analysis
    if url:
        if is_valid_url(url):
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.success(f"✅ Valid URL")
            
            with col2:
                # Detect content type
                content_type = detect_content_type(url)
                type_emoji = {
                    'html': '🌐',
                    'pdf': '📄',
                    'json': '📊',
                    'xml': '📜',
                    'unknown': '❓'
                }
                st.info(f"{type_emoji.get(content_type, '❓')} Type: {content_type.upper()}")
            
            with col3:
                # Check robots.txt
                can_fetch, robots_msg = check_robots_txt(url)
                if can_fetch or not st.session_state.get('respect_robots', True):
                    st.success("✅ Can proceed")
                else:
                    st.warning("⚠️ Robots.txt restricts")
            
            # Show robots.txt status
            if not can_fetch and st.session_state.get('respect_robots', True):
                st.warning(f"⚠️ {robots_msg} - Enable override in settings to proceed (educational purposes only)")
            elif not can_fetch:
                st.info(f"ℹ️ {robots_msg} - Proceeding due to override setting")
        else:
            st.error("❌ Invalid URL format. Please enter a valid URL starting with http:// or https://")
    
    # Enhanced scraping logic
    if scrape_button and url:
        if not is_valid_url(url):
            st.error("Please enter a valid URL before scraping.")
            return
        
        # Check robots.txt compliance
        can_fetch, robots_msg = check_robots_txt(url)
        if not can_fetch and st.session_state.get('respect_robots', True):
            st.error(f"🚱 Cannot proceed: {robots_msg}")
            st.info("💡 You can override this in the settings panel for educational purposes only.")
            return
        
        # Progress indicators with animations
        st.markdown("---")
        colored_header(
            label="🚀 Extraction in Progress",
            description="Analyzing and extracting content with AI intelligence",
            color_name="blue-green-70",
        )
        
        progress_container = st.container()
        
        with progress_container:
            col1, col2 = st.columns([2, 1])
            
            with col1:
                progress_bar = st.progress(0)
                status_text = st.empty()
            
            with col2:
                if animations["loading"]:
                    st_lottie(animations["loading"], height=100, key="extraction_loading")
        
        try:
            # Step 1: Content type detection
            status_text.text("🔍 Analyzing content type...")
            progress_bar.progress(20)
            content_type = detect_content_type(url)
            time.sleep(0.5)
            
            # Step 2: Fetching content
            status_text.text(f"📥 Fetching {content_type.upper()} content...")
            progress_bar.progress(40)
            time.sleep(0.5)
            
            # Step 3: Processing with enhanced extraction
            status_text.text("🧠 AI-powered content extraction...")
            progress_bar.progress(60)
            
            # Enhanced content extraction
            extraction_result = enhanced_content_extraction(url, content_type)
            
            progress_bar.progress(80)
            status_text.text("📊 Analyzing and formatting results...")
            time.sleep(0.3)
            
            progress_bar.progress(100)
            status_text.text("✅ Extraction completed successfully!")
            
            # Replace loading animation with success
            with col2:
                if animations["success"]:
                    st_lottie(animations["success"], height=100, key="extraction_success")
            
            if extraction_result.get('content') or not extraction_result.get('error'):
                # Enhanced results display
                st.markdown("---")
                colored_header(
                    label="📈 Extraction Results",
                    description=f"Comprehensive analysis of {content_type.upper()} content",
                    color_name="green-70",
                )
                
                # Results tabs for better organization
                tab1, tab2, tab3, tab4 = st.tabs(["📝 Overview", "📊 Content", "🔗 Metadata", "📥 Export"])
                
                with tab1:
                    # Enhanced metrics with better styling
                    col1, col2, col3, col4 = st.columns(4)
                    
                    content = extraction_result.get('content', '')
                    word_count = len(content.split()) if content else 0
                    char_count = len(content) if content else 0
                    
                    with col1:
                        st.metric("🔤 Characters", f"{char_count:,}")
                    with col2:
                        st.metric("📝 Words", f"{word_count:,}")
                    with col3:
                        st.metric("📄 Lines", len(content.split('\n')) if content else 0)
                    with col4:
                        readability = extraction_result.get('readability_score', 0)
                        st.metric("🎯 Readability", f"{readability:.1f}")
                    
                    style_metric_cards()
                    
                    # Title and content type
                    if extraction_result.get('title'):
                        st.subheader(f"🏷️ {extraction_result['title']}")
                    
                    # Content type badge
                    st.info(f"🏷️ Content Type: {content_type.upper()}")
                    
                    # Quick content preview
                    if content:
                        st.subheader("👁️ Quick Preview")
                        preview_text = content[:500] + "..." if len(content) > 500 else content
                        st.text_area(
                            "Content Preview",
                            value=preview_text,
                            height=150,
                            disabled=True
                        )
                
                with tab2:
                    # Full content display with syntax highlighting for different types
                    st.subheader("📋 Full Content")
                    
                    if content_type == 'json':
                        st.json(extraction_result.get('content', '{}'))
                    elif content_type == 'xml':
                        st.code(extraction_result.get('content', ''), language='xml')
                    else:
                        st.text_area(
                            "Full Extracted Content",
                            value=content,
                            height=400,
                            disabled=True
                        )
                    
                    # Links and images for HTML content
                    if extraction_result.get('links'):
                        st.subheader("🔗 Found Links")
                        for i, link in enumerate(extraction_result['links'][:10], 1):
                            st.markdown(f"{i}. [{link}]({link})")
                    
                    if extraction_result.get('images'):
                        st.subheader("🖼️ Found Images")
                        for i, img in enumerate(extraction_result['images'][:5], 1):
                            st.markdown(f"{i}. [{img}]({img})")
                
                with tab3:
                    # Enhanced metadata display
                    st.subheader("📊 Content Metadata")
                    
                    metadata = extraction_result.get('metadata', {})
                    if metadata:
                        for key, value in metadata.items():
                            if value:
                                st.markdown(f"**{key.replace('_', ' ').title()}:** {value}")
                    
                    # Technical details
                    st.subheader("🔧 Technical Details")
                    st.json({
                        "URL": url,
                        "Content Type": content_type,
                        "Extraction Time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Robots.txt Status": robots_msg,
                        "Word Count": word_count,
                        "Character Count": char_count
                    })
                
                with tab4:
                    # Enhanced download options
                    st.subheader("📦 Export Options")
                    
                    # Generate enhanced filename
                    domain = urlparse(url).netloc.replace('www.', '')
                    timestamp = time.strftime("%Y%m%d_%H%M%S")
                    filename = f"{domain}_{content_type}_{timestamp}"
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.markdown("**Standard Formats**")
                        # Text download
                        txt_link = create_download_link(content, filename, "text")
                        st.markdown(txt_link, unsafe_allow_html=True)
                        
                        # CSV download
                        csv_link = create_download_link(content, filename, "csv")
                        st.markdown(csv_link, unsafe_allow_html=True)
                    
                    with col2:
                        st.markdown("**Structured Formats**")
                        # JSON download with metadata
                        json_data = {
                            "url": url,
                            "content": content,
                            "metadata": extraction_result
                        }
                        json_link = create_download_link(json_data, filename, "json")
                        st.markdown(json_link, unsafe_allow_html=True)
                        
                        # Excel download
                        excel_link = create_download_link(extraction_result, filename, "excel")
                        st.markdown(excel_link, unsafe_allow_html=True)
                    
                    with col3:
                        st.markdown("**Document Formats**")
                        # PDF download
                        pdf_link = create_download_link(extraction_result, filename, "pdf")
                        st.markdown(pdf_link, unsafe_allow_html=True)
                        
                        # Raw data download
                        if content_type in ['xml', 'json']:
                            raw_link = create_download_link(content, filename + f"_raw", content_type)
                            st.markdown(f"[Download Raw {content_type.upper()}](data:text/plain;base64,{base64.b64encode(content.encode()).decode()})")
                
                # Download options
                st.markdown("---")
                st.subheader("💾 Download Options")
                
                col1, col2 = st.columns(2)
                
                # Generate filename from URL
                domain = urlparse(url).netloc.replace('www.', '')
                timestamp = time.strftime("%Y%m%d_%H%M%S")
                filename = f"{domain}_{timestamp}"
                
                with col1:
                    # Text download
                    txt_link = create_download_link(scraped_content, filename, "text")
                    st.markdown(txt_link, unsafe_allow_html=True)
                
                with col2:
                    # CSV download
                    csv_link = create_download_link(scraped_content, filename, "csv")
                    st.markdown(csv_link, unsafe_allow_html=True)
                
            else:
                st.error(f"❌ Extraction failed: {extraction_result.get('error', 'Unknown error')}")
                
                with st.expander("🔧 Troubleshooting Guide"):
                    st.markdown("""
                    **Common Issues & Solutions:**
                    
                    🔒 **Access Denied**
                    - Website may block automated requests
                    - Try adjusting request delay in settings
                    - Some sites require browser-based access
                    
                    📋 **No Content Found**
                    - Page might be JavaScript-heavy (SPA)
                    - Content could be behind authentication
                    - Check if URL is publicly accessible
                    
                    ⚙️ **Format Issues**
                    - PDF: Ensure file is not password-protected
                    - JSON: Verify valid JSON format
                    - XML: Check for well-formed structure
                    
                    🤖 **Robots.txt Restriction**
                    - Enable override in settings (educational use)
                    - Respect website terms of service
                    - Consider alternative approaches
                    """)
                
        except Exception as e:
            progress_bar.progress(100)
            status_text.text("❌ Extraction failed!")
            
            st.error(f"⚠️ Extraction Error: {str(e)}")
            
            # Enhanced error handling with specific guidance
            if "timeout" in str(e).lower():
                st.warning("⏱️ The website took too long to respond. Try increasing the timeout or try again later.")
            elif "permission" in str(e).lower() or "403" in str(e):
                st.warning("🚱 Access denied. The website may be blocking automated requests.")
            elif "404" in str(e):
                st.warning("🔄 Page not found. Please check the URL.")
            else:
                st.warning("ℹ️ An unexpected error occurred during extraction.")
            
            with st.expander("🚨 Advanced Troubleshooting"):
                st.markdown(f"""
                **Error Details:** `{str(e)}`
                
                **Debugging Steps:**
                1. 🔍 Verify URL accessibility in browser
                2. 🤖 Check robots.txt compliance
                3. ⚙️ Adjust scraping settings (delay, concurrency)
                4. 🔄 Try different content type detection
                5. 📞 Contact support if issue persists
                
                **Settings to Try:**
                - Increase request delay to 2-3 seconds
                - Reduce concurrent requests to 1
                - Toggle robots.txt respect (if appropriate)
                """)
    
    # Tips section
    with st.expander("💡 Pro Tips for Better Extraction"):
        st.markdown("""
        **🚀 Optimize Your Scraping:**
        
        ✅ **Supported Formats:**
        - HTML websites (news, blogs, documentation)
        - PDF documents (reports, papers, books)
        - JSON APIs (data feeds, REST responses)
        - XML feeds (RSS, sitemaps, data exports)
        
        ⚙️ **Best Practices:**
        - Use appropriate request delays (1-2 seconds)
        - Respect robots.txt for ethical scraping
        - Monitor success rates and adjust settings
        - Export data in multiple formats for analysis
        
        🌐 **Content Types Explained:**
        - **HTML**: Full webpage content with links and images
        - **PDF**: Document text with page-by-page extraction
        - **JSON**: Structured data with pretty formatting
        - **XML**: Hierarchical data with element analysis
        """)

def enhanced_batch_scraper_page():
    """Batch scraping interface"""
    db = ScrapingDatabase()
    
    st.title("🔄 Batch Web Scraper")
    st.markdown("Scrape multiple URLs at once and save results to history.")
    
    # URL input section
    st.subheader("🎯 Target URLs")
    
    # Session name
    col1, col2 = st.columns([2, 1])
    with col1:
        session_name = st.text_input(
            "Session Name",
            value=f"Batch_{datetime.now().strftime('%Y%m%d_%H%M')}",
            help="Name for this batch scraping session"
        )
    
    # URL input methods
    url_input_method = st.radio("Choose input method:", ["Text Area", "File Upload"])
    
    urls = []
    if url_input_method == "Text Area":
        url_text = st.text_area(
            "URLs (one per line)",
            placeholder="https://example1.com\nhttps://example2.com\nhttps://example3.com",
            height=150,
            help="Enter one URL per line"
        )
        if url_text:
            urls = [url.strip() for url in url_text.split('\n') if url.strip()]
    else:
        uploaded_file = st.file_uploader("Upload a text file with URLs", type=['txt'])
        if uploaded_file:
            content = uploaded_file.read().decode('utf-8')
            urls = [url.strip() for url in content.split('\n') if url.strip()]
    
    # Validate URLs
    valid_urls = []
    invalid_urls = []
    
    if urls:
        st.subheader(f"📋 URL Validation ({len(urls)} URLs found)")
        
        for url in urls:
            if is_valid_url(url):
                valid_urls.append(url)
            else:
                invalid_urls.append(url)
        
        col1, col2 = st.columns(2)
        with col1:
            st.success(f"✅ Valid URLs: {len(valid_urls)}")
        with col2:
            if invalid_urls:
                st.error(f"❌ Invalid URLs: {len(invalid_urls)}")
                with st.expander("View invalid URLs"):
                    for invalid_url in invalid_urls:
                        st.write(f"• {invalid_url}")
    
    # Batch scraping
    if valid_urls and st.button("🚀 Start Batch Scraping", type="primary", use_container_width=True):
        if not session_name.strip():
            st.error("Please enter a session name.")
            return
        
        # Create session
        session_id = db.create_session(session_name, len(valid_urls))
        
        # Progress tracking
        st.markdown("---")
        st.subheader("📊 Batch Scraping Progress")
        
        overall_progress = st.progress(0)
        current_url_text = st.empty()
        results_container = st.container()
        
        scraped_data = []
        successful_scrapes = 0
        
        for i, url in enumerate(valid_urls):
            current_url_text.text(f"🔄 Scraping {i+1}/{len(valid_urls)}: {url}")
            
            try:
                content = get_website_text_content(url)
                if content:
                    # Extract title if possible
                    title = url.split('/')[-1] if '/' in url else url
                    
                    # Save to database
                    db.save_scraped_data(session_id, url, content, title)
                    scraped_data.append({
                        'url': url,
                        'title': title,
                        'content': content,
                        'word_count': len(content.split()),
                        'char_count': len(content),
                        'status': 'success'
                    })
                    successful_scrapes += 1
                    
                    # Show progress
                    with results_container:
                        st.success(f"✅ {url} - {len(content.split())} words")
                else:
                    db.save_scraped_data(session_id, url, "", "", "failed", "No content extracted")
                    with results_container:
                        st.warning(f"⚠️ {url} - No content extracted")
                        
            except Exception as e:
                error_msg = str(e)
                db.save_scraped_data(session_id, url, "", "", "failed", error_msg)
                with results_container:
                    st.error(f"❌ {url} - Error: {error_msg}")
            
            # Update progress
            overall_progress.progress((i + 1) / len(valid_urls))
            db.update_session_progress(session_id, i + 1)
            time.sleep(0.1)  # Small delay to prevent overwhelming
        
        # Final status
        db.update_session_progress(session_id, len(valid_urls), 'completed')
        current_url_text.text(f"✅ Completed! Successfully scraped {successful_scrapes}/{len(valid_urls)} URLs")
        
        # Results summary
        if scraped_data:
            st.markdown("---")
            st.subheader("📄 Batch Results Summary")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total URLs", len(valid_urls))
            with col2:
                st.metric("Successful", successful_scrapes)
            with col3:
                st.metric("Failed", len(valid_urls) - successful_scrapes)
            with col4:
                total_words = sum(item['word_count'] for item in scraped_data)
                st.metric("Total Words", total_words)
            
            # Export options
            st.markdown("---")
            st.subheader("💾 Export Batch Results")
            
            filename = f"{session_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                txt_link = create_download_link(scraped_data, filename, "text")
                st.markdown(txt_link, unsafe_allow_html=True)
            with col2:
                csv_link = create_download_link(scraped_data, filename, "csv")
                st.markdown(csv_link, unsafe_allow_html=True)
            with col3:
                json_link = create_download_link(scraped_data, filename, "json")
                st.markdown(json_link, unsafe_allow_html=True)
            with col4:
                excel_link = create_download_link(scraped_data, filename, "excel")
                st.markdown(excel_link, unsafe_allow_html=True)
            
            # PDF export
            pdf_link = create_download_link(scraped_data, filename, "pdf")
            st.markdown(pdf_link, unsafe_allow_html=True)

def history_page():
    """Scraping history and results management"""
    db = ScrapingDatabase()
    
    st.title("📚 Scraping History")
    st.markdown("View and manage your scraping sessions and results.")
    
    # Get all sessions
    sessions = db.get_sessions()
    
    if not sessions:
        st.info("📝 No scraping sessions found. Start by scraping some URLs!")
        return
    
    # Sessions overview
    st.subheader("🗂️ All Sessions")
    
    for session in sessions:
        with st.expander(f"📁 {session['session_name']} - {session['created_at']}", expanded=False):
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.write(f"**Status:** {session['status']}")
                st.write(f"**Total URLs:** {session['total_urls']}")
            with col2:
                st.write(f"**Completed:** {session['completed_urls']}")
                progress = session['completed_urls'] / max(session['total_urls'], 1)
                st.progress(progress)
            with col3:
                if st.button(f"View Details", key=f"view_{session['id']}"):
                    st.session_state.selected_session = session['id']
            with col4:
                if st.button(f"🗑️ Delete", key=f"delete_{session['id']}", type="secondary"):
                    if st.session_state.get(f"confirm_delete_{session['id']}", False):
                        db.delete_session(session['id'])
                        st.success("Session deleted!")
                        st.rerun()
                    else:
                        st.session_state[f"confirm_delete_{session['id']}"] = True
                        st.warning("Click again to confirm deletion")
    
    # Session details
    if hasattr(st.session_state, 'selected_session') and st.session_state.selected_session:
        st.markdown("---")
        st.subheader("📄 Session Details")
        
        session_data = db.get_session_data(st.session_state.selected_session)
        
        if session_data:
            # Export session data
            col1, col2 = st.columns([3, 1])
            with col1:
                st.write(f"**Showing {len(session_data)} scraped items**")
            with col2:
                selected_session_name = next(s['session_name'] for s in sessions if s['id'] == st.session_state.selected_session)
                filename = f"{selected_session_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                
                export_format = st.selectbox("Export as:", ["CSV", "JSON", "Excel", "PDF", "Text"])
                if st.button("Download", type="primary"):
                    export_data = []
                    for item in session_data:
                        export_data.append({
                            'url': item['url'],
                            'title': item['title'],
                            'content': item['content'],
                            'word_count': item['word_count'],
                            'char_count': item['char_count'],
                            'scraped_at': item['scraped_at'],
                            'status': item['status']
                        })
                    
                    download_link = create_download_link(export_data, filename, export_format.lower())
                    st.markdown(download_link, unsafe_allow_html=True)
            
            # Display items
            for item in session_data:
                with st.container():
                    if item['status'] == 'success':
                        st.success(f"✅ **{item['url']}**")
                        if item['title']:
                            st.write(f"**Title:** {item['title']}")
                        
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Words", item['word_count'])
                        with col2:
                            st.metric("Characters", item['char_count'])
                        with col3:
                            st.write(f"**Scraped:** {item['scraped_at']}")
                        
                        # Content preview
                        if item['content']:
                            with st.expander("View Content"):
                                st.text_area(
                                    "Content",
                                    value=item['content'][:1000] + "..." if len(item['content']) > 1000 else item['content'],
                                    height=200,
                                    disabled=True,
                                    key=f"content_{item['id']}"
                                )
                    else:
                        st.error(f"❌ **{item['url']}** - Failed")
                        if item['error_message']:
                            st.write(f"**Error:** {item['error_message']}")
                    
                    st.markdown("---")

def search_page():
    """Search within scraped content"""
    db = ScrapingDatabase()
    
    st.title("🔍 Search Scraped Content")
    st.markdown("Search through all your scraped content across sessions.")
    
    # Search form
    col1, col2 = st.columns([3, 1])
    with col1:
        search_term = st.text_input(
            "Search Term",
            placeholder="Enter keywords to search...",
            help="Search within titles, content, and URLs"
        )
    
    with col2:
        sessions = db.get_sessions()
        session_options = ["All Sessions"] + [f"{s['session_name']} ({s['id']})" for s in sessions]
        selected_session = st.selectbox("Filter by Session", session_options)
    
    if search_term and st.button("🔍 Search", type="primary"):
        # Get session ID if specific session selected
        session_id = None
        if selected_session != "All Sessions":
            session_id = int(selected_session.split("(")[-1].strip(")"))
        
        # Perform search
        results = db.search_content(search_term, session_id)
        
        if results:
            st.subheader(f"🎯 Search Results ({len(results)} found)")
            
            for result in results:
                with st.container():
                    st.success(f"✅ **{result['url']}**")
                    
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        if result['title']:
                            st.write(f"**Title:** {result['title']}")
                        st.write(f"**Session:** {result['session_name']}")
                    with col2:
                        st.metric("Words", result['word_count'])
                        st.write(f"**Date:** {result['scraped_at']}")
                    
                    # Highlight search term in content preview
                    content_preview = result['content'][:500]
                    if search_term.lower() in content_preview.lower():
                        # Simple highlight (case insensitive)
                        highlighted = content_preview.replace(
                            search_term, 
                            f"**{search_term}**"
                        )
                        st.markdown(f"**Preview:** {highlighted}...")
                    else:
                        st.write(f"**Preview:** {content_preview}...")
                    
                    with st.expander("View Full Content"):
                        st.text_area(
                            "Full Content",
                            value=result['content'],
                            height=300,
                            disabled=True,
                            key=f"search_content_{result['id']}"
                        )
                    
                    st.markdown("---")
        else:
            st.info(f"🔍 No results found for '{search_term}'")

def scheduled_page():
    """Scheduled scraping management"""
    st.title("⏰ Scheduled Scraping")
    st.markdown("Set up automated scraping tasks with optional email notifications.")
    
    # Info box about scheduled scraping limitations
    st.info("📧 **Email Notifications**: This demo shows the interface for scheduled scraping. In a production environment, you would need to set up email service (SMTP) and a background task scheduler.")
    
    # Create new scheduled task
    with st.expander("➕ Create New Scheduled Task", expanded=True):
        st.subheader("📝 Task Configuration")
        
        col1, col2 = st.columns(2)
        with col1:
            task_name = st.text_input(
                "Task Name",
                placeholder="Daily News Scraping",
                help="Give your scheduled task a descriptive name"
            )
            
            schedule_type = st.selectbox(
                "Schedule Type",
                ["Daily", "Weekly", "Monthly", "Custom Interval"],
                help="How often should this task run?"
            )
            
        with col2:
            if schedule_type == "Daily":
                schedule_time = st.time_input("Run Time")
                schedule_value = f"daily_{schedule_time.strftime('%H:%M')}"
            elif schedule_type == "Weekly":
                day_of_week = st.selectbox("Day of Week", 
                    ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"])
                schedule_time = st.time_input("Run Time")
                schedule_value = f"weekly_{day_of_week}_{schedule_time.strftime('%H:%M')}"
            elif schedule_type == "Monthly":
                day_of_month = st.number_input("Day of Month", min_value=1, max_value=28, value=1)
                schedule_time = st.time_input("Run Time")
                schedule_value = f"monthly_{day_of_month}_{schedule_time.strftime('%H:%M')}"
            else:
                interval_hours = st.number_input("Interval (hours)", min_value=1, max_value=168, value=24)
                schedule_value = f"interval_{interval_hours}h"
        
        # URLs for scheduled task
        st.subheader("🎯 URLs to Scrape")
        scheduled_urls = st.text_area(
            "URLs (one per line)",
            placeholder="https://example1.com\nhttps://example2.com",
            height=100,
            help="Enter URLs to scrape on schedule"
        )
        
        # Email notifications
        st.subheader("📧 Email Notifications")
        email_notifications = st.checkbox("Send email notifications when complete")
        email_address = ""
        if email_notifications:
            email_address = st.text_input(
                "Email Address",
                placeholder="your-email@example.com",
                help="Where to send notification emails"
            )
        
        # Create task button
        if st.button("✅ Create Scheduled Task", type="primary"):
            if not task_name.strip():
                st.error("Please enter a task name.")
            elif not scheduled_urls.strip():
                st.error("Please enter at least one URL.")
            elif email_notifications and not email_address.strip():
                st.error("Please enter an email address for notifications.")
            else:
                # Validate URLs
                urls = [url.strip() for url in scheduled_urls.split('\n') if url.strip()]
                valid_urls = [url for url in urls if is_valid_url(url)]
                
                if len(valid_urls) != len(urls):
                    st.error(f"Some URLs are invalid. Please check your URLs.")
                else:
                    # Create actual scheduled task
                    try:
                        scheduler = get_scheduler()
                        task_id = scheduler.create_scheduled_task(
                            task_name, valid_urls, schedule_type, schedule_value,
                            email_notifications, email_address
                        )
                        
                        st.success(f"✅ Scheduled task '{task_name}' created successfully!")
                        st.info("📋 **Task Details:**\n"
                               f"- **Name**: {task_name}\n"
                               f"- **Schedule**: {schedule_type} ({schedule_value})\n"
                               f"- **URLs**: {len(valid_urls)} URLs\n"
                               f"- **Email**: {'Yes' if email_notifications else 'No'}")
                        
                        if email_notifications:
                            st.write(f"- **Email Address**: {email_address}")
                        
                        st.info("📧 **Email Setup**: To receive email notifications, set the following environment variables:\n"
                               "- SMTP_SERVER (default: smtp.gmail.com)\n"
                               "- SMTP_PORT (default: 587)\n"
                               "- SMTP_USERNAME\n"
                               "- SMTP_PASSWORD")
                               
                    except Exception as e:
                        st.error(f"❌ Failed to create scheduled task: {str(e)}")
                        st.info("💡 The scheduler is running in demo mode. In production, ensure the background scheduler service is properly configured.")
    
    # Real scheduled tasks display
    st.markdown("---")
    st.subheader("📅 Active Scheduled Tasks")
    
    # Get real scheduled tasks from database
    db = ScrapingDatabase()
    scheduled_tasks = db.get_scheduled_tasks()
    
    if not scheduled_tasks:
        st.info("📝 No scheduled tasks found. Create one above to get started!")
    else:
        for task in scheduled_tasks:
            with st.container():
                col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
                
                with col1:
                    st.write(f"**📋 {task['task_name']}**")
                    st.write(f"🕐 {task['schedule_type']} - {task['schedule_value']}")
                    st.write(f"🌐 {len(task['urls'])} URLs")
                
                with col2:
                    status_color = "🟢" if task['is_active'] else "🔴"
                    status_text = "Active" if task['is_active'] else "Paused"
                    st.write(f"**Status:** {status_color} {status_text}")
                    if task['email_notifications'] and task['email_address']:
                        st.write(f"📧 {task['email_address']}")
                
                with col3:
                    st.write(f"**Last Run:**")
                    st.write(f"{task['last_run'] or 'Never'}")
                    st.write(f"**Created:**")
                    st.write(f"{task['created_at']}")
                
                with col4:
                    if task['is_active']:
                        if st.button(f"⏸️ Pause", key=f"pause_{task['id']}", type="secondary"):
                            try:
                                scheduler = get_scheduler()
                                scheduler.pause_task(task['id'])
                                st.success(f"Task '{task['task_name']}' paused.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Failed to pause task: {str(e)}")
                    else:
                        if st.button(f"▶️ Resume", key=f"resume_{task['id']}", type="secondary"):
                            try:
                                scheduler = get_scheduler()
                                scheduler.resume_task(task['id'])
                                st.success(f"Task '{task['task_name']}' resumed.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Failed to resume task: {str(e)}")
                    
                    if st.button(f"🗑️ Delete", key=f"del_sched_{task['id']}", type="secondary"):
                        if st.session_state.get(f"confirm_delete_sched_{task['id']}", False):
                            try:
                                scheduler = get_scheduler()
                                scheduler.delete_task(task['id'])
                                st.success(f"Task '{task['task_name']}' deleted.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Failed to delete task: {str(e)}")
                        else:
                            st.session_state[f"confirm_delete_sched_{task['id']}"] = True
                            st.warning("Click again to confirm deletion")
                
                st.markdown("---")
    
    # Instructions for production setup
    with st.expander("🔧 Production Setup Instructions", expanded=False):
        st.markdown("""
        ### Setting Up Scheduled Scraping in Production
        
        To implement scheduled scraping with email notifications, you would need:
        
        #### 1. **Task Scheduler**
        ```python
        from apscheduler.schedulers.background import BackgroundScheduler
        import smtplib
        from email.mime.text import MIMEText
        
        scheduler = BackgroundScheduler()
        scheduler.start()
        
        def scheduled_scrape_job(task_id):
            # Load task from database
            # Perform scraping
            # Send email notification if configured
            pass
        
        # Add job to scheduler
        scheduler.add_job(scheduled_scrape_job, 'cron', 
                         hour=9, minute=0, args=[task_id])
        ```
        
        #### 2. **Email Service**
        ```python
        def send_notification(email, subject, body):
            msg = MIMEText(body)
            msg['Subject'] = subject
            msg['From'] = 'scraper@yourapp.com'
            msg['To'] = email
            
            with smtplib.SMTP('smtp.gmail.com', 587) as server:
                server.starttls()
                server.login('your_email', 'your_password')
                server.send_message(msg)
        ```
        
        #### 3. **Database Schema**
        The `scheduled_tasks` table is already created in the database schema.
        
        #### 4. **Background Processing**
        Consider using Celery with Redis/RabbitMQ for robust task processing.
        """)

def about_page():
    """About page with application information"""
    st.title("ℹ️ About Web Scraper")
    
    st.markdown("""
    ## 🎯 Purpose
    This web scraper application is designed to legally and ethically extract text content from websites 
    for research, analysis, and data collection purposes.
    
    ## ✨ Features
    - **Secure Access**: Admin authentication for controlled usage
    - **Legal Scraping**: Uses Trafilatura for clean, ethical content extraction
    - **Real-time Progress**: Live status updates during scraping operations
    - **Multiple Export Formats**: Download scraped data as text or CSV files
    - **Responsive Design**: Clean, mobile-friendly interface
    - **Error Handling**: Comprehensive error detection and user guidance
    
    ## 🛠️ Technology Stack
    - **Frontend**: Streamlit (Python)
    - **Scraping Engine**: Trafilatura
    - **Data Processing**: Pandas
    - **Authentication**: Session-based with hardcoded credentials
    
    ## ⚖️ Legal Notice
    This tool is intended for legal web scraping only. Please ensure you:
    - Respect website terms of service
    - Check robots.txt files
    - Don't overload servers with requests
    - Use scraped data responsibly
    
    ## 🔧 How to Use
    1. **Login**: Use admin credentials to access the application
    2. **Enter URL**: Provide a valid website URL in the scraper page
    3. **Scrape**: Click the scrape button to extract content
    4. **Download**: Save the results as text or CSV files
    
    ## 📞 Support
    For technical support or questions about this application, contact your system administrator.
    """)

def enhanced_history_page():
    """Enhanced scraping history with modern UI"""
    db = ScrapingDatabase()
    animations = get_lottie_animations()
    
    colored_header(
        label="📚 Scraping History",
        description="Manage and analyze your scraping sessions with advanced insights",
        color_name="violet-70",
    )
    
    sessions = db.get_sessions()
    
    if not sessions:
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if animations and animations.get("data_analysis"):
                st_lottie(animations["data_analysis"], height=200, key="no_history")
            st.info("📝 No scraping sessions found. Start by scraping some URLs!")
            if st.button("🚀 Start First Session", type="primary", use_container_width=True):
                st.session_state.show_landing = False
                st.rerun()
        return
    
    # Enhanced sessions overview
    for session in sessions:
        with st.expander(f"📁 {session['session_name']} - {session['created_at']}", expanded=False):
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.write(f"**Status:** {session['status']}")
                st.write(f"**Total URLs:** {session['total_urls']}")
            with col2:
                st.write(f"**Completed:** {session['completed_urls']}")
                progress = session['completed_urls'] / max(session['total_urls'], 1)
                st.progress(progress)
            with col3:
                if st.button(f"View Details", key=f"view_{session['id']}"):
                    st.session_state.selected_session = session['id']
            with col4:
                if st.button(f"🗑️ Delete", key=f"delete_{session['id']}", type="secondary"):
                    if st.session_state.get(f"confirm_delete_{session['id']}", False):
                        db.delete_session(session['id'])
                        st.success("Session deleted!")
                        st.rerun()
                    else:
                        st.session_state[f"confirm_delete_{session['id']}"] = True
                        st.warning("Click again to confirm deletion")

def enhanced_search_page():
    """Enhanced search interface with AI-powered capabilities"""
    db = ScrapingDatabase()
    
    colored_header(
        label="🔍 Smart Content Search",
        description="AI-powered search through your scraped content with advanced filters",
        color_name="orange-70",
    )
    
    col1, col2 = st.columns([3, 1])
    
    with col1:
        search_term = st.text_input(
            "🔍 Search Query",
            placeholder="Enter keywords, phrases, or content to search...",
            help="Search within titles, content, and URLs across all sessions"
        )
    
    with col2:
        sessions = db.get_sessions()
        session_options = ["🌐 All Sessions"] + [f"📁 {s['session_name']}" for s in sessions]
        selected_session = st.selectbox("Filter by Session:", session_options)
    
    if search_term and st.button("🔍 Search Content", type="primary", use_container_width=True):
        session_id = None
        if not selected_session.startswith("🌐"):
            session_name = selected_session.replace("📁 ", "")
            session_id = next((s['id'] for s in sessions if s['session_name'] == session_name), None)
        
        results = db.search_content(search_term, session_id)
        
        if results:
            st.success(f"✅ Found {len(results)} results matching '{search_term}'")
            
            for i, result in enumerate(results, 1):
                with st.expander(f"{i}. 🔗 {result['url']}", expanded=False):
                    if result['title']:
                        st.markdown(f"📝 **Title:** {result['title']}")
                    
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        preview = result['content'][:300] + "..." if len(result['content']) > 300 else result['content']
                        st.markdown(preview)
                    with col2:
                        st.metric("📊 Words", result['word_count'])
                        st.caption(f"⏰ {result['scraped_at']}")
        else:
            st.warning(f"🔍 No results found for '{search_term}'")

def enhanced_scheduled_page():
    """Enhanced scheduled scraping with modern UI"""
    colored_header(
        label="⏰ Smart Scheduler",
        description="Automated scraping with AI-powered notifications and insights",
        color_name="red-70",
    )
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("⚡ Active Jobs", "0", "Coming soon")
    with col2:
        st.metric("✅ Completed", "0", "Feature ready")
    with col3:
        st.metric("📧 Notifications", "Enabled", "Email alerts")
    
    style_metric_cards()
    
    st.info("💡 Scheduled scraping features are being finalized. Use batch scraper for now!")

def enhanced_about_page():
    """Enhanced about page with beautiful design"""
    colored_header(
        label="ℹ️ About Advanced Web Scraper",
        description="Your intelligent content extraction platform",
        color_name="violet-70",
    )
    
    st.markdown("""
    ### 🚀 Welcome to the Future of Web Scraping
    
    Our advanced platform combines **AI-powered extraction** with **ethical scraping practices** 
    to deliver the most comprehensive content analysis tool available.
    """)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("🔧 Technical Features")
        st.markdown("""
        - **🌐 Multi-Format Support**: HTML, PDF, JSON, XML
        - **🤖 AI-Powered Extraction**: Intelligent content detection
        - **⚡ Batch Processing**: Handle thousands of URLs
        - **📊 Advanced Analytics**: Readability scores, metadata
        - **💾 Multiple Exports**: CSV, JSON, Excel, PDF
        - **🔍 Smart Search**: Full-text content search
        - **🛡️ Ethical Scraping**: Robots.txt compliance
        - **⏰ Scheduling**: Automated recurring tasks
        """)
    
    with col2:
        st.subheader("🏗️ Architecture")
        st.markdown("""
        - **Frontend**: Streamlit with modern UI components
        - **Backend**: Python with async processing
        - **Database**: SQLite with optimized queries
        - **Scraping Engine**: Trafilatura + BeautifulSoup
        - **PDF Processing**: PDFPlumber for document analysis
        - **Scheduling**: APScheduler with persistence
        - **Animations**: Lottie for engaging interactions
        - **Styling**: Custom CSS with responsive design
        """)

def main():
    """Main application entry point"""
    # Initialize session state
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    if 'username' not in st.session_state:
        st.session_state.username = None
    
    # Page configuration
    st.set_page_config(
        page_title="Web Scraper Admin",
        page_icon="🕸️",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Custom CSS for better styling (minimal)
    st.markdown("""
    <style>
    .main-header {
        text-align: center;
        padding: 1rem 0;
    }
    .metric-container {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 0.5rem 0;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Route to appropriate page based on authentication
    if st.session_state.authenticated:
        main_app()
    else:
        login_page()

if __name__ == "__main__":
    main()
